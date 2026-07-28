"""
LAM+ Core QC — qc_avaatech.py
Streamlit Frontend for Avaatech XRF Core Scanner QC

Igor Oliveira, Andre Belem, F.R.I.D.A.Y. / LAM+

Uso:
    streamlit run qc_avaatech.py
"""

import io
import os
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill

from qc_core import (
    CORE_DEPTH_COL,
    DEFAULT_ENERGY,
    DEPTH_COL,
    ENERGY_PARAMETERS,
    QF_INDETERMINATE,
    QF_PLOT_ORDER,
    add_pointwise_flag_notes,
    check_file,
    read_workbook,
    run_qc,
)
from i18n import TEXTS, DEFAULT_LANG

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PNG = os.path.join(BASE_DIR, "assets", "lamplus_logo.png")

QF_COLORS = {0: "#2ca02c", 1: "#ff7f0e", 2: "#d62728", 3: "#7f0000", QF_INDETERMINATE: "#7f7f7f"}

# Rótulo de exibição curto para a variável efetivamente plotada em
# plot_rolling — Throughput é o fallback quando a energia não mede nenhuma
# variável espectral (50 kV, ver ENERGY_PARAMETERS/plot_rolling).
ROLLING_VAR_LABELS = {
    "Rh-La-Inc Area": "Rh-Lα-Inc",
    "Rh-Ka-Inc Area": "Rh-Kα-Inc",
    "Throughput": "Throughput",
}

# ============================================================
# FIGURAS
# ============================================================

def plot_throughput(rep0, T, depth_col=DEPTH_COL):
    fig, ax = plt.subplots(figsize=(10, 3))
    ax.plot(rep0[depth_col], rep0["Throughput"], color="#2c7bb6", lw=0.8)
    ax.axhline(rep0["Throughput"].median(), color="gray", ls="--", lw=0.7, label=T["plot_throughput_median"])
    ax.set_xlabel(T["depth_axis"])
    ax.set_ylabel("Throughput")
    ax.set_title(T["plot_throughput_title"])
    ax.legend(fontsize=8)
    fig.tight_layout()
    return fig


def plot_rh(rep0, T, depth_col=DEPTH_COL, energy=DEFAULT_ENERGY):
    """
    QC2/QC3 — Coherent/Incoherent Scatter. Colunas físicas dependem da
    energia (Rh-La Area/Rh-La-Inc Area em 10 kV, Rh-Ka-Coh Area/
    Rh-Ka-Inc Area em 30 kV — ver ENERGY_PARAMETERS). Retorna None quando a
    energia não mede nenhum dos dois (50 kV) — o chamador mostra uma
    mensagem informativa nesse caso em vez do gráfico.
    """
    cfg = ENERGY_PARAMETERS[energy]
    cols = [c for c in (cfg["coherent"], cfg["incoherent"]) if c is not None and c in rep0.columns]
    if not cols:
        return None
    titles = [T["plot_rh1_title"], T["plot_rh2_title"]][: len(cols)]
    fig, axes = plt.subplots(len(cols), 1, figsize=(10, 2.5 * len(cols)), sharex=True)
    if len(cols) == 1:
        axes = [axes]
    for ax, col, title in zip(axes, cols, titles):
        ax.plot(rep0[depth_col], rep0[col], lw=0.8)
        ax.set_ylabel(col)
        ax.set_title(title)
    axes[-1].set_xlabel(T["depth_axis"])
    fig.tight_layout()
    return fig


def plot_rolling(rep0, T, depth_col=DEPTH_COL, energy=DEFAULT_ENERGY):
    """
    QC4 — Rolling QC. Plota a mesma variável que alimenta Rolling_z por
    padrão em compute_scores (incoerente da energia; Throughput quando a
    energia não mede nenhuma variável espectral — 50 kV, ver
    ENERGY_PARAMETERS). Para 10 kV reproduz exatamente o gráfico anterior a
    esta mudança (Rh-Lα-Inc).
    """
    cfg = ENERGY_PARAMETERS[energy]
    var = cfg["incoherent"]
    if var is None or f"{var}_rolling" not in rep0.columns:
        var = cfg["throughput"]
    var_label = ROLLING_VAR_LABELS.get(var, var)

    fig, ax = plt.subplots(figsize=(10, 3))
    ax.plot(rep0[depth_col], rep0[var], lw=0.8, label=T["plot_rolling_original"])
    ax.plot(rep0[depth_col], rep0[f"{var}_rolling"], lw=1.2, ls="--", label=T["plot_rolling_mean"])
    ax.set_xlabel(T["depth_axis"])
    ax.set_ylabel(var)
    ax.set_title(T["plot_rolling_title"].format(var=var_label))
    ax.legend(fontsize=8)
    fig.tight_layout()
    return fig


def plot_pca(rep0, p95, p99, T):
    fig, ax = plt.subplots(figsize=(6, 5))
    sc = ax.scatter(rep0["PC1"], rep0["PC2"], c=rep0["Mahalanobis"], cmap="RdYlGn_r", s=15, alpha=0.7)
    plt.colorbar(sc, ax=ax, label=T["plot_pca_colorbar"])
    ax.set_xlabel("PC1")
    ax.set_ylabel("PC2")
    ax.set_title(T["plot_pca_title"])
    fig.tight_layout()
    return fig


def plot_qi(rep0, T, depth_col=DEPTH_COL):
    c = rep0["QF"].map(QF_COLORS).fillna("#aaaaaa")
    qf_plot_pos = rep0["QF"].map(QF_PLOT_ORDER)
    fig, axes = plt.subplots(2, 1, figsize=(10, 5), sharex=True)
    axes[0].bar(rep0[depth_col], rep0["QI"].fillna(0), color=c.values, width=8)
    axes[0].axhline(80, color="gray", ls="--", lw=0.7)
    axes[0].axhline(40, color="red", ls="--", lw=0.7)
    axes[0].set_ylabel(T["plot_qi_ylabel"])
    axes[0].set_title(T["plot_qi_title"])
    axes[1].bar(rep0[depth_col], qf_plot_pos, color=c.values, width=8)
    axes[1].set_yticks(list(QF_PLOT_ORDER.values()))
    axes[1].set_yticklabels(T["plot_qf_labels"])
    axes[1].set_xlabel(T["depth_axis"])
    axes[1].set_ylabel(T["plot_qf_ylabel"])
    fig.tight_layout()
    return fig


def plot_replicas(rep0, T, depth_col=DEPTH_COL):
    fig, ax = plt.subplots(figsize=(10, 3))
    valid = rep0.dropna(subset=["Mean_RPD"])
    ax.bar(valid[depth_col], valid["Mean_RPD"], width=8, color="#9467bd", alpha=0.8)
    ax.set_xlabel(T["depth_axis"])
    ax.set_ylabel(T["plot_replicas_ylabel"])
    ax.set_title(T["plot_replicas_title"])
    fig.tight_layout()
    return fig


# ============================================================
# EXPORT
# ============================================================

# Coloração por célula (protocolo v4.2, DEVELOPMENT.md — reports.py do
# apêndice): verde=OK/QF0/"NO", amarelo=Atenção/QF1/"WARNING",
# vermelho=Suspeito-Rejeitado/QF2-QF3/"CRITICAL"/"YES".
QC_FILL_GREEN = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
QC_FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
QC_FILL_RED = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")


def _qc_cell_fill(value, is_qf_column):
    """
    Cor para uma célula de coluna QC, conforme protocolo v4.2. Os códigos
    numéricos 0/1/2/3 só se aplicam à coluna QF: em colunas numéricas
    contínuas (scores, z-scores, RPD) esses mesmos valores não indicam
    severidade e coincidir com eles não deve disparar coloração.
    """
    if value in ("OK", "NO") or (is_qf_column and value == 0):
        return QC_FILL_GREEN
    if value == "WARNING" or (is_qf_column and value == 1):
        return QC_FILL_YELLOW
    if value in ("CRITICAL", "YES") or (is_qf_column and value in (2, 3)):
        return QC_FILL_RED
    return None


def to_excel_bytes(sheet_results):
    """
    Serializa o(s) resultado(s) de QC para .xlsx — uma aba por energia
    processada, preservando o nome original da aba (`sheet_name`) e as
    colunas originais do Avaatech; aplica coloração verde/amarelo/vermelho
    (via openpyxl.PatternFill) célula a célula, restrita às colunas
    adicionadas pelo pipeline QC (ausentes das colunas originais daquela
    aba) — nunca nas colunas originais.

    sheet_results: list de dicts com "sheet_name", "rep0" (DataFrame com QC)
        e "df_raw" (DataFrame bruto da aba, para saber quais colunas são
        originais).
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        used_names = set()
        for result in sheet_results:
            df = result["rep0"]
            original_columns = set(result["df_raw"].columns)

            # Nomes de aba do Excel têm limite de 31 caracteres e devem ser
            # únicos no workbook; sheet_name original raramente colide, mas
            # garantimos unicidade em vez de deixar openpyxl levantar erro.
            sheet_name = str(result["sheet_name"])[:31]
            base_name, suffix = sheet_name, 1
            while sheet_name in used_names:
                suffix += 1
                sheet_name = f"{base_name[:28]}_{suffix}"
            used_names.add(sheet_name)

            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name in original_columns:
                    continue
                is_qf_column = col_name == "QF"
                for row_idx, value in enumerate(df[col_name], start=2):
                    fill = _qc_cell_fill(value, is_qf_column)
                    if fill is not None:
                        ws.cell(row=row_idx, column=col_idx).fill = fill

    return buf.getvalue()


# ============================================================
# INTERFACE
# ============================================================

st.set_page_config(
    page_title=TEXTS[DEFAULT_LANG]["page_title"],
    page_icon=None,
    layout="wide",
)

LANG_OPTIONS = {TEXTS[l]["language_name"]: l for l in TEXTS}
lang_label = st.sidebar.selectbox(
    TEXTS[DEFAULT_LANG]["language_selector_label"],
    options=list(LANG_OPTIONS.keys()),
    index=0,
)
lang = LANG_OPTIONS[lang_label]
T = TEXTS[lang]

DEPTH_DISPLAY_OPTIONS = {
    T["depth_display_composite"]: DEPTH_COL,
    T["depth_display_core"]: CORE_DEPTH_COL,
}
depth_display_label = st.sidebar.selectbox(
    T["depth_display_label"],
    options=list(DEPTH_DISPLAY_OPTIONS.keys()),
    index=0,
    help=T["depth_display_help"],
)
DEPTH_DISPLAY_COL = DEPTH_DISPLAY_OPTIONS[depth_display_label]

strict_missing_data = st.sidebar.checkbox(
    T["strict_missing_label"],
    value=True,
    help=T["strict_missing_help"],
)
combine_rolling_vars = st.sidebar.checkbox(
    T["combine_rolling_label"],
    value=False,
    help=T["combine_rolling_help"],
)
include_pca_in_qf = st.sidebar.checkbox(
    T["include_pca_qf_label"],
    value=False,
    help=T["include_pca_qf_help"],
)
use_count_mode = st.sidebar.checkbox(
    T["use_count_mode_label"],
    value=False,
    help=T["use_count_mode_help"],
)
use_rolling_persistence = st.sidebar.checkbox(
    T["rolling_persistence_label"],
    value=False,
    help=T["rolling_persistence_help"],
)

col_title, col_logo = st.columns([5, 1])
with col_title:
    st.title(T["app_title"])
    st.caption(T["app_caption"])
with col_logo:
    if os.path.isfile(LOGO_PNG):
        st.image(LOGO_PNG, width=100)

uploaded = st.file_uploader(T["upload_label"], type=["xlsx"])

if uploaded is None:
    st.info(T["upload_info"])
    st.stop()

# Leitura do workbook inteiro — detecta a energia (10 kV/30 kV/50 kV) de
# cada aba pelo nome (protocolo v4.2, ver qc_core.detect_energy). Um
# arquivo de aba única cujo nome não segue essa convenção assume 10 kV
# (DEFAULT_ENERGY) — mantém compatibilidade com exports mais antigos.
try:
    sheets, skipped_sheets = read_workbook(uploaded)
except Exception as e:
    st.error(T["read_error"].format(error=e))
    st.stop()

if not sheets:
    st.error(T["no_sheets_error"])
    st.stop()

st.success(T["load_success"].format(
    rows=sum(s["df"].shape[0] for s in sheets),
    cols=sum(s["df"].shape[1] for s in sheets),
))
st.caption(T["workbook_sheets_info"].format(
    sheets=", ".join(f"{s['sheet_name']} ({s['energy']})" for s in sheets)
))
if skipped_sheets:
    st.warning(T["workbook_skipped_warning"].format(sheets=", ".join(skipped_sheets)))

# Check de consistência + QC — cada aba é validada e processada
# independentemente; uma aba com erro não impede as demais de aparecer.
sheet_results = []
with st.spinner(T["qc_spinner"]):
    for sheet in sheets:
        sheet_name, energy, df_raw = sheet["sheet_name"], sheet["energy"], sheet["df"]

        errors, warnings = check_file(df_raw, lang=lang, energy=energy)
        if errors:
            for e in errors:
                st.error(T["error_prefix"].format(msg=f"[{sheet_name}] {e}"))
            continue
        for w in warnings:
            st.warning(T["warning_prefix"].format(msg=f"[{sheet_name}] {w}"))

        try:
            rep0, p95, p99, pca_elements = run_qc(
                df_raw,
                strict_missing_data=strict_missing_data,
                combine_rolling_vars=combine_rolling_vars,
                include_pca_in_qf=include_pca_in_qf,
                use_count_mode=use_count_mode,
                use_rolling_persistence=use_rolling_persistence,
                energy=energy,
            )
            rep0 = add_pointwise_flag_notes(rep0, lang=lang)
        except Exception as e:
            st.error(T["qc_error"].format(error=f"[{sheet_name}] {e}"))
            continue

        sheet_results.append({
            "sheet_name": sheet_name,
            "energy": energy,
            "df_raw": df_raw,
            "rep0": rep0,
            "p95": p95,
            "p99": p99,
            "pca_elements": pca_elements,
        })

if not sheet_results:
    st.stop()

# Seletor de aba/energia — só exibido quando há mais de uma aba processada
# com sucesso, para não mudar a experiência do arquivo de aba única.
if len(sheet_results) > 1:
    sheet_options = {
        f"{r['sheet_name']} ({r['energy']})": i for i, r in enumerate(sheet_results)
    }
    selected_label = st.selectbox(T["energy_sheet_selector_label"], options=list(sheet_options.keys()))
    selected = sheet_results[sheet_options[selected_label]]
else:
    selected = sheet_results[0]

rep0 = selected["rep0"]
p95 = selected["p95"]
p99 = selected["p99"]
energy = selected["energy"]

# Resumo
st.subheader(T["summary_header"])
col1, col2, col3, col4, col5 = st.columns(5)
qf_counts = rep0["QF"].value_counts().sort_index()
col1.metric(T["metric_measurements"], len(rep0))
col2.metric(T["metric_mean_qi"], f"{rep0['QI'].mean():.1f}")
col3.metric(T["metric_qf_ok"], qf_counts.get(0, 0))
col4.metric(T["metric_qf_rejected"], qf_counts.get(3, 0))
col5.metric(T["metric_qf_indeterminate"], qf_counts.get(QF_INDETERMINATE, 0))

# Figuras
st.subheader(T["diagnostics_header"])

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    T["tab_throughput"], T["tab_rh"], T["tab_rolling"],
    T["tab_replicas"], T["tab_pca"], T["tab_qi"],
])

with tab1:
    st.pyplot(plot_throughput(rep0, T, depth_col=DEPTH_DISPLAY_COL))

with tab2:
    fig_rh = plot_rh(rep0, T, depth_col=DEPTH_DISPLAY_COL, energy=energy)
    if fig_rh is not None:
        st.pyplot(fig_rh)
    else:
        st.info(T["rh_not_applicable_info"])

with tab3:
    st.pyplot(plot_rolling(rep0, T, depth_col=DEPTH_DISPLAY_COL, energy=energy))

with tab4:
    st.pyplot(plot_replicas(rep0, T, depth_col=DEPTH_DISPLAY_COL))

with tab5:
    if rep0["Mahalanobis"].notna().any():
        st.pyplot(plot_pca(rep0, p95, p99, T))
    else:
        st.info(T["pca_skipped_info"])

with tab6:
    st.pyplot(plot_qi(rep0, T, depth_col=DEPTH_DISPLAY_COL))

# Tabela
st.subheader(T["data_header"])
display_cols = [DEPTH_DISPLAY_COL] + [c for c in rep0.columns if c != DEPTH_DISPLAY_COL]
st.dataframe(rep0[display_cols], use_container_width=True)

# Download — workbook com todas as abas processadas (não só a selecionada),
# cada uma preservando seu nome e colunas originais.
st.download_button(
    label=T["download_label"],
    data=to_excel_bytes(sheet_results),
    file_name="LAM_CoreQC_Output.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
