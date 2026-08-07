"""
Testes da Etapa 6 do plano-de-acao.md (saídas e interface mínima, em
qc_reports.py: build_excel_report, build_summary, format_summary_text).

Casos cobertos:
    - colunas originais preservadas intactas (valores e ordem);
    - colunas QC presentes e com os valores corretos ao final da aba;
    - coloração condicional restrita às colunas QC -- nunca nas colunas
      originais, mesmo quando uma coluna original tem valores que
      coincidem textualmente com estados QC (ex. "OK"/"ALERT"/"CRITICAL");
    - aba "Flagged_Intervals" contém só as linhas Review="YES" de todas as
      abas, identificadas por Sheet/Energy;
    - resumo textual contém os campos esperados (arquivo, por energia:
      n(Rep0)/distribuição de QF/contagem ALERT-CRITICAL por módulo, lista
      de profundidades com Review=YES).

Usa DataFrames sintéticos já no formato de saída de qc_core.run_qc (não
recalcula QC1-QC5 -- isso já é coberto por test_qc_core.py/
test_qc_integrate.py), para testar qc_reports.py isoladamente.

Não depende de LEGACY/ (ver conftest.py).
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from qc_config import QCState, QualityFlag
from qc_reports import (
    FLAGGED_INTERVALS_SHEET,
    QC_EXPORT_COLUMNS,
    build_excel_report,
    build_summary,
    format_summary_text,
)

# Lê os locales diretamente do JSON em vez de `import i18n`: este módulo
# roda no mesmo processo que test_qc_core_vs_legacy.py, que carrega
# LEGACY/qc_core.py (e o `i18n.py` dele) via manipulação de sys.path/
# sys.modules sob o mesmo nome "i18n". Um `import i18n` aqui deixaria a V2
# ou a LEGACY em sys.modules["i18n"] dependendo da ordem de execução dos
# testes, quebrando um dos dois lados (ver comentário em
# qc_avaatech._select_language). Ler o JSON direto evita a colisão.
_LOCALES_DIR = Path(__file__).resolve().parents[1] / "locales"


def _load_strings(lang: str) -> dict[str, str]:
    with (_LOCALES_DIR / f"{lang}.json").open(encoding="utf-8") as fh:
        return json.load(fh)


OK = QCState.OK
ALERT = QCState.ALERT
CRITICAL = QCState.CRITICAL
NA = QCState.NOT_APPLICABLE


def make_sheet_result(sheet_name: str = "10kV", energy: str = "10kV") -> dict:
    """
    Aba sintética de 3 medições já no formato pós run_qc: 1 linha QF0 (tudo
    OK), 1 linha QF1 (QC1 ALERT) e 1 linha QF2 (QC1 CRITICAL). Inclui uma
    coluna original "Status" com valores literais "OK"/"ALERT"/"CRITICAL"
    -- serve para provar que a coloração nunca toca colunas originais,
    mesmo quando o valor coincide textualmente com um estado QC.
    """
    df_raw = pd.DataFrame(
        {
            "Spectrum": ["s0", "s1", "s2"],
            "CoreDepth": [10.0, 20.0, 30.0],
            "CompositeDepth (mm)": [10.0, 20.0, 30.0],
            "Replicate Nr Count": ["Rep0", "Rep0", "Rep0"],
            "Throughput": [1000.0, 1001.0, 1002.0],
            "Status": ["OK", "ALERT", "CRITICAL"],
        }
    )

    rep0 = df_raw.copy()
    rep0["QC1_State"] = [OK, ALERT, CRITICAL]
    rep0["QC2_State"] = [OK, OK, OK]
    rep0["QC3_State"] = [OK, OK, OK]
    rep0["QC4_State"] = [OK, OK, OK]
    rep0["QC5_State"] = [NA, NA, NA]
    rep0["QF_Cause"] = [None, "throughput", "throughput"]
    rep0["QF_Evidence"] = ["", "", ""]
    rep0["QF"] = [QualityFlag.QF0, QualityFlag.QF1, QualityFlag.QF2]
    rep0["Review"] = ["NO", "NO", "YES"]

    return {"sheet_name": sheet_name, "energy": energy, "rep0": rep0, "df_raw": df_raw}


def _load_sheet(xlsx_bytes: bytes, sheet_name: str):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    return wb[sheet_name]


def _header(ws) -> list[str]:
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _fill_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fgColor is None or fill.fgColor.rgb in (None, "00000000"):
        return None
    return fill.fgColor.rgb


# ============================================================
# build_excel_report -- colunas originais
# ============================================================


def test_original_columns_preserved_intact_and_in_order():
    sheet_result = make_sheet_result()
    df_raw = sheet_result["df_raw"]

    xlsx_bytes = build_excel_report([sheet_result])
    ws = _load_sheet(xlsx_bytes, "10kV")
    header = _header(ws)

    assert header[: len(df_raw.columns)] == list(df_raw.columns)

    for col_idx, col_name in enumerate(df_raw.columns, start=1):
        expected = df_raw[col_name].tolist()
        actual = [row[col_idx - 1] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert actual == expected, f"coluna original alterada: {col_name}"


# ============================================================
# build_excel_report -- colunas QC
# ============================================================


def test_qc_columns_present_and_correct_at_the_end():
    sheet_result = make_sheet_result()
    df_raw = sheet_result["df_raw"]

    xlsx_bytes = build_excel_report([sheet_result])
    ws = _load_sheet(xlsx_bytes, "10kV")
    header = _header(ws)

    assert header[len(df_raw.columns):] == list(QC_EXPORT_COLUMNS)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    col_idx = {name: i for i, name in enumerate(header)}

    assert [r[col_idx["QC1_State"]] for r in rows] == ["OK", "ALERT", "CRITICAL"]
    assert [r[col_idx["QualityFlag"]] for r in rows] == ["QF0", "QF1", "QF2"]
    assert [r[col_idx["Review"]] for r in rows] == ["NO", "NO", "YES"]
    assert [r[col_idx["QF_Cause"]] for r in rows] == [None, "throughput", "throughput"]


def test_missing_required_keys_raise_keyerror():
    incomplete = {"sheet_name": "10kV", "energy": "10kV", "rep0": pd.DataFrame()}
    with pytest.raises(KeyError):
        build_excel_report([incomplete])


# ============================================================
# build_excel_report -- coloração restrita às colunas QC
# ============================================================


def test_coloring_never_touches_original_columns():
    sheet_result = make_sheet_result()
    xlsx_bytes = build_excel_report([sheet_result])
    ws = _load_sheet(xlsx_bytes, "10kV")
    header = _header(ws)
    status_col = header.index("Status") + 1

    # "Status" é original e tem valores literais "OK"/"ALERT"/"CRITICAL" --
    # mesmo assim nenhuma célula deve ser colorida.
    for row_idx in range(2, 5):
        assert _fill_rgb(ws.cell(row=row_idx, column=status_col)) is None


def test_coloring_applied_correctly_on_qc_columns():
    sheet_result = make_sheet_result()
    xlsx_bytes = build_excel_report([sheet_result])
    ws = _load_sheet(xlsx_bytes, "10kV")
    header = _header(ws)
    qc1_col = header.index("QC1_State") + 1
    qf_col = header.index("QualityFlag") + 1
    review_col = header.index("Review") + 1

    # linha 2 = OK/QF0/NO -> verde; linha 3 = ALERT/QF1/NO -> amarelo (QC1),
    # verde (Review); linha 4 = CRITICAL/QF2/YES -> vermelho.
    assert _fill_rgb(ws.cell(row=2, column=qc1_col)) == "00C6EFCE"
    assert _fill_rgb(ws.cell(row=3, column=qc1_col)) == "00FFF2CC"
    assert _fill_rgb(ws.cell(row=4, column=qc1_col)) == "00F4CCCC"

    assert _fill_rgb(ws.cell(row=2, column=qf_col)) == "00C6EFCE"
    assert _fill_rgb(ws.cell(row=3, column=qf_col)) == "00FFF2CC"
    assert _fill_rgb(ws.cell(row=4, column=qf_col)) == "00F4CCCC"

    assert _fill_rgb(ws.cell(row=2, column=review_col)) == "00C6EFCE"
    assert _fill_rgb(ws.cell(row=3, column=review_col)) == "00C6EFCE"
    assert _fill_rgb(ws.cell(row=4, column=review_col)) == "00F4CCCC"


# ============================================================
# build_excel_report -- múltiplas abas / nomes únicos
# ============================================================


def test_one_output_sheet_per_input_sheet_preserving_names():
    result_10 = make_sheet_result(sheet_name="10kV", energy="10kV")
    result_30 = make_sheet_result(sheet_name="30kV", energy="30kV")

    xlsx_bytes = build_excel_report([result_10, result_30])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    assert "10kV" in wb.sheetnames
    assert "30kV" in wb.sheetnames


# ============================================================
# build_excel_report -- aba Flagged_Intervals
# ============================================================


def test_flagged_intervals_sheet_contains_only_review_yes_rows_across_sheets():
    result_10 = make_sheet_result(sheet_name="10kV", energy="10kV")
    result_30 = make_sheet_result(sheet_name="30kV", energy="30kV")

    xlsx_bytes = build_excel_report([result_10, result_30])
    ws = _load_sheet(xlsx_bytes, FLAGGED_INTERVALS_SHEET)
    header = _header(ws)

    assert header[:2] == ["Sheet", "Energy"]
    assert "Review" in header

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    review_idx = header.index("Review")
    sheet_idx = header.index("Sheet")

    # cada aba de entrada contribui com exatamente 1 linha Review=YES
    assert len(rows) == 2
    assert all(r[review_idx] == "YES" for r in rows)
    assert {r[sheet_idx] for r in rows} == {"10kV", "30kV"}


def test_flagged_intervals_sheet_empty_when_no_rows_flagged():
    sheet_result = make_sheet_result()
    sheet_result["rep0"]["Review"] = ["NO", "NO", "NO"]

    xlsx_bytes = build_excel_report([sheet_result])
    ws = _load_sheet(xlsx_bytes, FLAGGED_INTERVALS_SHEET)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows == []


# ============================================================
# build_summary / format_summary_text
# ============================================================


def test_build_summary_aggregates_per_energy():
    sheet_result = make_sheet_result(sheet_name="10kV", energy="10kV")
    summary = build_summary([sheet_result], file_name="arquivo.xlsx")

    assert summary["file_name"] == "arquivo.xlsx"
    stats = summary["energies"]["10kV"]
    assert stats["n_measurements"] == 3
    assert stats["qf_distribution"] == {
        "QF0": 1, "QF1": 1, "QF2": 1, "QF3": 0, "INDETERMINATE": 0,
    }
    assert stats["module_counts"]["QC1_State"] == {"ALERT": 1, "CRITICAL": 1}
    assert stats["module_counts"]["QC2_State"] == {"ALERT": 0, "CRITICAL": 0}

    assert len(summary["flagged"]) == 1
    flagged = summary["flagged"][0]
    assert flagged["sheet_name"] == "10kV"
    assert flagged["energy"] == "10kV"
    assert flagged["depth"] == 30.0
    assert flagged["cause"] == "throughput"


def test_build_summary_does_not_mutate_input():
    sheet_result = make_sheet_result()
    original_columns = list(sheet_result["rep0"].columns)
    build_summary([sheet_result])
    assert list(sheet_result["rep0"].columns) == original_columns


def test_format_summary_text_contains_expected_fields_pt():
    sheet_result = make_sheet_result(sheet_name="10kV", energy="10kV")
    summary = build_summary([sheet_result], file_name="meu_arquivo.xlsx")
    text = format_summary_text(summary, _load_strings("pt"))

    assert "Arquivo: meu_arquivo.xlsx" in text
    assert "Energia 10kV:" in text
    assert "n(Rep0) = 3" in text
    assert "QF0: 1 | QF1: 1 | QF2: 1 | QF3: 0 | INDETERMINATE: 0" in text
    assert "ALERT/CRITICAL por módulo:" in text
    assert "QC1_State: ALERT=1 CRITICAL=1" in text
    assert "Profundidades com Review=YES:" in text
    assert "depth=30.0" in text
    assert "causa: throughput" in text


def test_format_summary_text_reports_no_flagged_rows_pt():
    sheet_result = make_sheet_result()
    sheet_result["rep0"]["Review"] = ["NO", "NO", "NO"]
    summary = build_summary([sheet_result], file_name="arquivo.xlsx")
    text = format_summary_text(summary, _load_strings("pt"))

    assert "Profundidades com Review=YES:" in text
    assert "(nenhuma)" in text


def test_format_summary_text_contains_expected_fields_en():
    sheet_result = make_sheet_result(sheet_name="10kV", energy="10kV")
    summary = build_summary([sheet_result], file_name="my_file.xlsx")
    text = format_summary_text(summary, _load_strings("en"))

    assert "File: my_file.xlsx" in text
    assert "Energy 10kV:" in text
    assert "n(Rep0) = 3" in text
    assert "QF0: 1 | QF1: 1 | QF2: 1 | QF3: 0 | INDETERMINATE: 0" in text
    assert "ALERT/CRITICAL by module:" in text
    assert "QC1_State: ALERT=1 CRITICAL=1" in text
    assert "Depths with Review=YES:" in text
    assert "depth=30.0" in text
    assert "cause: throughput" in text

    # nenhum texto em português deve vazar para o resumo em inglês.
    assert "Arquivo" not in text
    assert "Energia" not in text
    assert "módulo" not in text
    assert "Profundidades" not in text
    assert "causa" not in text


def test_format_summary_text_reports_no_flagged_rows_en():
    sheet_result = make_sheet_result()
    sheet_result["rep0"]["Review"] = ["NO", "NO", "NO"]
    summary = build_summary([sheet_result], file_name="file.xlsx")
    text = format_summary_text(summary, _load_strings("en"))

    assert "Depths with Review=YES:" in text
    assert "(none)" in text


def test_format_summary_text_no_file_name_uses_translated_placeholder():
    sheet_result = make_sheet_result()
    summary = build_summary([sheet_result])

    assert "File: (no name)" in format_summary_text(summary, _load_strings("en"))
    assert "Arquivo: (sem nome)" in format_summary_text(summary, _load_strings("pt"))
