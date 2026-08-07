"""
LAM+ Core QC V2 — qc_reports.py

Responsabilidade: exportação dos resultados em Excel e geração de resumo
simples. Não contém cálculo científico (isso é responsabilidade de
qc_core.py) nem lógica de interface (qc_avaatech.py). Não importa nada de
LEGACY/.
"""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill

from qc_config import QCState, QualityFlag
from qc_io import resolve_depth_column

# ============================================================
# CONTRATO DE SAÍDA
# ============================================================

# Colunas QC adicionadas ao final de cada aba exportada, nesta ordem exata
# (Etapa 6 do plano-de-acao.md). Não inclui colunas intermediárias do
# cálculo (z-scores, Mean_RPD, QC_Module_States, contagens) -- essas ficam
# só no DataFrame interno de qc_core.run_qc; a saída para o usuário é
# deliberadamente reduzida a este conjunto (DEVELOPMENT.md 4.1: saída
# direta e rastreável, sem excesso de colunas).
QC_EXPORT_COLUMNS = (
    "QC1_State",
    "QC2_State",
    "QC3_State",
    "QC4_State",
    "QC5_State",
    "QF_Cause",
    "QF_Evidence",
    "QualityFlag",
    "Review",
)

FLAGGED_INTERVALS_SHEET = "Flagged_Intervals"

_STATE_COLUMNS = ("QC1_State", "QC2_State", "QC3_State", "QC4_State", "QC5_State")

# Coloração condicional (protocolo v4.2): verde=OK/QF0/NO,
# amarelo=ALERT/QF1, vermelho=CRITICAL/QF2/QF3/YES. NOT_APPLICABLE e
# INDETERMINATE ficam sem cor -- são neutros, não "bons" nem "ruins" (ver
# qc_config.QCState).
QC_FILL_GREEN = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
QC_FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
QC_FILL_RED = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")

_CELL_FILL_BY_VALUE = {
    "OK": QC_FILL_GREEN,
    "NO": QC_FILL_GREEN,
    "QF0": QC_FILL_GREEN,
    "ALERT": QC_FILL_YELLOW,
    "QF1": QC_FILL_YELLOW,
    "CRITICAL": QC_FILL_RED,
    "YES": QC_FILL_RED,
    "QF2": QC_FILL_RED,
    "QF3": QC_FILL_RED,
}


def format_quality_flag(qf: QualityFlag) -> str:
    """Rótulo textual curto de um QualityFlag: 'QF0'..'QF3' ou 'INDETERMINATE'."""
    if qf == QualityFlag.INDETERMINATE:
        return "INDETERMINATE"
    return f"QF{int(qf)}"


def _sheet_export_frame(sheet_result: dict) -> pd.DataFrame:
    """
    Monta o DataFrame de exportação de uma aba: colunas originais do
    instrumento intactas e na ordem original, seguidas das colunas QC de
    QC_EXPORT_COLUMNS.

    Requer que sheet_result["rep0"] já tenha passado por qc_core.run_qc
    (contém QC1_State..QC5_State, QF, QF_Cause, QF_Evidence, Review) e que
    sheet_result["df_raw"] seja a aba bruta correspondente (usada só para
    saber quais colunas são originais, nunca suas linhas).
    """
    rep0 = sheet_result["rep0"]
    original_columns = [c for c in sheet_result["df_raw"].columns if c in rep0.columns]

    out = rep0[original_columns].copy()
    for col in QC_EXPORT_COLUMNS:
        if col == "QualityFlag":
            out[col] = rep0["QF"].map(format_quality_flag)
        else:
            out[col] = rep0[col].to_numpy()
    return out


def _unique_sheet_name(name: str, used_names: set[str]) -> str:
    """
    Garante um nome de aba válido para o Excel: até 31 caracteres e único
    dentro do workbook (openpyxl levanta erro em duplicatas).
    """
    sheet_name = str(name)[:31]
    base_name, suffix = sheet_name, 1
    while sheet_name in used_names:
        suffix += 1
        sheet_name = f"{base_name[:28]}_{suffix}"
    used_names.add(sheet_name)
    return sheet_name


def _apply_qc_coloring(ws, df: pd.DataFrame, original_columns: set[str]) -> None:
    """
    Aplica a coloração condicional célula a célula, restrita às colunas de
    `df` que NÃO estão em `original_columns` -- nunca colore uma coluna
    original do instrumento, mesmo que seu valor coincida por acaso com uma
    das strings de estado (ex. uma coluna de texto do Avaatech com o valor
    literal "OK").
    """
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in original_columns:
            continue
        for row_idx, value in enumerate(df[col_name], start=2):
            fill = _CELL_FILL_BY_VALUE.get(value)
            if fill is not None:
                ws.cell(row=row_idx, column=col_idx).fill = fill


def build_excel_report(sheet_results: list[dict]) -> bytes:
    """
    Serializa os resultados de todas as abas processadas com sucesso num
    único workbook .xlsx para download.

    Entrada:
        sheet_results: list[dict], um item por aba processada, cada um
            contendo "sheet_name" (str), "energy" (str), "rep0" (DataFrame
            já processado por qc_core.run_qc) e "df_raw" (DataFrame bruto
            da aba, usado só para identificar quais colunas são originais).

    Saída:
        Bytes do arquivo .xlsx pronto para download/gravação.

    Contrato:
        - Uma aba de saída por aba de entrada, preservando o nome original
          e as colunas originais do instrumento sem alteração.
        - Colunas QC (QC_EXPORT_COLUMNS) são acrescentadas ao final, nunca
          substituem colunas originais.
        - Coloração condicional restrita às colunas QC acrescentadas --
          nunca nas colunas originais.
        - Aba extra "Flagged_Intervals" com todas as linhas Review="YES" de
          todas as abas, identificadas por "Sheet"/"Energy".
        - Não achata nem combina abas de energias diferentes.
        - Não modifica `sheet_results` nem os DataFrames nele contidos.
    """
    # Monta todas as abas de saída (e valida as chaves obrigatórias de cada
    # sheet_result) antes de abrir o ExcelWriter -- assim um dict malformado
    # levanta KeyError diretamente, em vez de ser mascarado por um erro de
    # "workbook sem abas" no __exit__ do writer.
    used_names: set[str] = {FLAGGED_INTERVALS_SHEET}
    prepared: list[tuple[str, str, str, pd.DataFrame, set[str]]] = []
    for sheet_result in sheet_results:
        export_df = _sheet_export_frame(sheet_result)
        original_columns = set(sheet_result["df_raw"].columns)
        output_sheet_name = _unique_sheet_name(sheet_result["sheet_name"], used_names)
        prepared.append(
            (
                output_sheet_name,
                sheet_result["sheet_name"],
                sheet_result["energy"],
                export_df,
                original_columns,
            )
        )

    flagged_frames: list[pd.DataFrame] = []
    for output_sheet_name, original_sheet_name, energy, export_df, _ in prepared:
        flagged = export_df[export_df["Review"] == "YES"].copy()
        if not flagged.empty:
            flagged.insert(0, "Energy", energy)
            flagged.insert(0, "Sheet", original_sheet_name)
            flagged_frames.append(flagged)

    flagged_df = (
        pd.concat(flagged_frames, ignore_index=True)
        if flagged_frames
        else pd.DataFrame(columns=["Sheet", "Energy", *QC_EXPORT_COLUMNS])
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for output_sheet_name, _, _, export_df, original_columns in prepared:
            export_df.to_excel(writer, index=False, sheet_name=output_sheet_name)
            _apply_qc_coloring(writer.sheets[output_sheet_name], export_df, original_columns)

        flagged_df.to_excel(writer, index=False, sheet_name=FLAGGED_INTERVALS_SHEET)
        _apply_qc_coloring(
            writer.sheets[FLAGGED_INTERVALS_SHEET], flagged_df, {"Sheet", "Energy"}
        )

    return buf.getvalue()


def build_summary(sheet_results: list[dict], file_name: str = "") -> dict[str, Any]:
    """
    Gera um resumo simples dos resultados processados: por energia, número
    de medições (Rep0), distribuição de QF0-QF3/INDETERMINATE e contagem de
    ALERT/CRITICAL por módulo; mais a lista de profundidades com
    Review="YES", causa principal e evidências.

    Entrada:
        sheet_results: mesma estrutura de build_excel_report.
        file_name: nome do arquivo de origem, incluído no resumo.

    Saída:
        dict com chaves "file_name", "energies" (dict energia -> métricas)
        e "flagged" (lista de dicts sheet_name/energy/depth/cause/evidence).
        Use format_summary_text para uma versão textual pronta para exibir
        ou baixar.

    Contrato:
        - Não modifica `sheet_results` nem os DataFrames nele contidos.
        - QualityFlag.INDETERMINATE é contado separadamente de QF0-QF3,
          nunca somado a QF0.
        - Abas com a mesma energia têm suas medições agregadas juntas.
    """
    energies: dict[str, dict[str, Any]] = {}
    flagged: list[dict[str, Any]] = []

    for sheet_result in sheet_results:
        rep0 = sheet_result["rep0"]
        energy = sheet_result["energy"]
        sheet_name = sheet_result["sheet_name"]

        bucket = energies.setdefault(
            energy,
            {
                "n_measurements": 0,
                "qf_distribution": {
                    "QF0": 0, "QF1": 0, "QF2": 0, "QF3": 0, "INDETERMINATE": 0,
                },
                "module_counts": {
                    col: {"ALERT": 0, "CRITICAL": 0} for col in _STATE_COLUMNS
                },
            },
        )

        bucket["n_measurements"] += len(rep0)
        for qf in rep0["QF"]:
            bucket["qf_distribution"][format_quality_flag(qf)] += 1
        for col in _STATE_COLUMNS:
            bucket["module_counts"][col]["ALERT"] += int((rep0[col] == QCState.ALERT).sum())
            bucket["module_counts"][col]["CRITICAL"] += int(
                (rep0[col] == QCState.CRITICAL).sum()
            )

        try:
            depth_col, _ = resolve_depth_column(rep0)
        except ValueError:
            depth_col = None

        for _, row in rep0[rep0["Review"] == "YES"].iterrows():
            flagged.append(
                {
                    "sheet_name": sheet_name,
                    "energy": energy,
                    "depth": row[depth_col] if depth_col is not None else None,
                    "cause": row["QF_Cause"],
                    "evidence": row["QF_Evidence"],
                }
            )

    return {"file_name": file_name, "energies": energies, "flagged": flagged}


def format_summary_text(summary: dict[str, Any], strings: dict[str, str]) -> str:
    """
    Renderiza o dict de build_summary como texto simples, pronto para
    exibição ou download (.txt).

    Entrada:
        summary: saída de build_summary.
        strings: dict de strings do i18n (ver i18n.load) -- todos os
            rótulos e mensagens do texto gerado vêm das chaves
            "summary_*" deste dict, nunca hardcoded aqui.

    Contrato:
        - O idioma do texto gerado é inteiramente determinado por
          `strings`: passar i18n.load("en") produz um resumo inteiramente
          em inglês; i18n.load("pt"), inteiramente em português.
    """
    file_label = summary.get("file_name") or strings["summary_no_file_name"]
    lines = [strings["summary_file_label"].format(file=file_label), ""]

    for energy, stats in summary["energies"].items():
        qf = stats["qf_distribution"]
        lines.append(strings["summary_energy_label"].format(energy=energy))
        lines.append("  " + strings["summary_n_measurements"].format(n=stats["n_measurements"]))
        lines.append("  " + strings["summary_qf_distribution"].format(**qf))
        lines.append("  " + strings["summary_module_header"])
        for module, counts in stats["module_counts"].items():
            lines.append(
                "    "
                + strings["summary_module_line"].format(
                    module=module, alert=counts["ALERT"], critical=counts["CRITICAL"]
                )
            )
        lines.append("")

    lines.append(strings["summary_flagged_header"])
    if summary["flagged"]:
        for item in summary["flagged"]:
            depth = item["depth"]
            depth_str = f"{depth:.1f}" if isinstance(depth, (int, float)) else str(depth)
            evidence = item["evidence"] or "-"
            lines.append(
                "  "
                + strings["summary_flagged_line"].format(
                    energy=item["energy"],
                    sheet=item["sheet_name"],
                    depth=depth_str,
                    cause=item["cause"],
                    evidence=evidence,
                )
            )
    else:
        lines.append("  " + strings["summary_flagged_none"])

    return "\n".join(lines)
